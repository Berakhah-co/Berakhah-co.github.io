from bs4 import BeautifulSoup
import pandas as pd
import os
import sys  # Importamos 'sys' para salir del script de forma controlada
import openpyxl
from openpyxl.styles import Alignment

# Función para extraer los datos del HTML (sin cambios)
def extraer_productos(html):
    soup = BeautifulSoup(html, 'html.parser')
    productos = []
    
    productos_html = soup.find_all('div', class_='producto')
    
    for producto in productos_html:
        categoria = ""
        nombre = ""
        sku = ""
        descripcion = ""
        precio = "Sin precio"

        titulo = producto.find('h3', class_='titulo-producto')
        if titulo:
            nombre = titulo.text.strip()

        if nombre:
            sku = nombre.split()[-1]

        parrafos = producto.find_all('p')
        textos_descripcion = []
        for p in parrafos:
            if 'precio' not in p.get('class', []):
                textos_descripcion.append(p.get_text(separator=' ', strip=True))
        
        descripcion = "\n".join(textos_descripcion)

        precio_tag = producto.find('p', class_='precio')
        if precio_tag:
            precio_texto = precio_tag.text.strip().replace('$', '').replace('.', '').replace(',', '')
            if precio_texto.isdigit():
                precio = int(precio_texto)

        nombre_lower = nombre.lower()
        if nombre_lower.startswith('collar'):
            categoria = "Moda > Joyería > Collares y Cadenas"
        elif nombre_lower.startswith('pulsera'):
            categoria = "Moda > Joyería > Pulseras y Tobilleras"
        elif nombre_lower.startswith('anillo'):
            categoria = "Moda > Joyería > Anillos"
        elif nombre_lower.startswith('zarcillo') or nombre_lower.startswith('arete'):
            categoria = "Moda > Joyería > Aretes y Earcuff"
        else:
            categoria = "Categoría Ejemplo"

        marca = ""
        ean = ""
        es_pesable = "NO"
        es_preempaquetado = "NO"
        cantidad = 1
        unidad_medida = "Und (unidades)"
        
        productos.append([
            categoria, nombre, sku, marca, ean, descripcion,
            precio, es_pesable, es_preempaquetado, cantidad, unidad_medida
        ])

    df = pd.DataFrame(productos, columns=[
        'Categoría', 'Nombre', 'SKU', 'Marca (opcional)', 'EAN (opcional)', 'Descripción',
        'Precio', '¿Es pesable?', '¿Es preempaquetado?', 'Cantidad', 'Unidad de medida'
    ])
    
    return df

# Función para eliminar duplicados (sin cambios)
def eliminar_duplicados(ws, columna):
    valores_vistos = set()
    filas_a_eliminar = []

    for row_idx in range(2, ws.max_row + 1):
        celda_valor = ws[f"{columna}{row_idx}"].value
        if celda_valor in valores_vistos:
            filas_a_eliminar.append(row_idx)
        else:
            if celda_valor is not None:
                valores_vistos.add(celda_valor)

    for row_idx in reversed(filas_a_eliminar):
        ws.delete_rows(row_idx)

# --- INICIO DE LA SECCIÓN MODIFICADA ---

# Función para leer el archivo HTML y generar el archivo Excel
def generar_excel_desde_html(html_path, excel_path):
    try:
        with open(html_path, 'r', encoding='utf-8') as file:
            html = file.read()
        
        df = extraer_productos(html)
        
        # Intentamos guardar el archivo Excel
        df.to_excel(excel_path, index=False, engine='openpyxl')
        
        print(f"Excel generado exitosamente: {excel_path}")
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column == 'F':
                ws.column_dimensions[column].width = 50
            else:
                adjusted_width = (max_length + 2) if max_length < 40 else 40
                ws.column_dimensions[column].width = adjusted_width

        eliminar_duplicados(ws, 'C')

        wb.save(excel_path)
        
        os.startfile(excel_path)

    # Si ocurre un error de permisos, mostramos un mensaje amigable
    except PermissionError:
        print("\n------------------------------------------------------------------")
        print(f"ERROR: Permiso denegado para escribir en el archivo '{os.path.basename(excel_path)}'.")
        print("SOLUCIÓN: Por favor, CIERRA el archivo de Excel y vuelve a ejecutar el script.")
        print("------------------------------------------------------------------")
        sys.exit(1) # Salimos del script para evitar más errores
        
    # Si no se encuentra el HTML, también damos un mensaje claro
    except FileNotFoundError:
        print(f"ERROR: No se pudo encontrar el archivo '{os.path.basename(html_path)}'.")
        print("SOLUCIÓN: Asegúrate de que 'index.html' esté en la misma carpeta que el script.")
        sys.exit(1)

# --- FIN DE LA SECCIÓN MODIFICADA ---

# Rutas de los archivos (sin cambios)
script_dir = os.path.dirname(os.path.abspath(__file__))
html_path = os.path.join(script_dir, 'index.html')
excel_path = os.path.join(script_dir, 'productos.xlsx')

# Ejecutar la función principal
generar_excel_desde_html(html_path, excel_path)